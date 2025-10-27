#!/usr/bin/env python3
"""Command line helper that converts PDFs to Excel using UniversalBankExtractor."""

import logging
import os
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .camelot_utils import extract_tables_with_camelot
from .universal_extractor import UniversalBankExtractor


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler("pdf2xls.log")],
)

logger = logging.getLogger("pdf2xls")


def highlight_problem_rows(excel_file: str) -> None:
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if "observaciones" not in headers:
            logger.warning("Column 'observaciones' not found in %s", excel_file)
            return

        column_index = headers.index("observaciones") + 1
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row_idx in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=column_index).value
            if value and str(value).strip():
                for col_idx in range(1, sheet.max_column + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = yellow

        workbook.save(excel_file)
    except Exception as exc:  # pragma: no cover - auxiliary script
        logger.warning("Could not highlight rows in %s: %s", excel_file, exc)


def format_excel_output(dataframe: pd.DataFrame, output_file: Path) -> None:
    formatted = dataframe.copy()
    
    # 🔹 Formatear montos (debitos, creditos, saldo)
    for column in ["debitos", "creditos", "saldo", "debito", "credito"]:
        if column in formatted.columns:
            formatted[column] = formatted[column].apply(
                lambda value: (
                    f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    if pd.notna(value) and value != ""
                    else "0,00"
                )
            )

    # 🔹 CRÍTICO: Mantener fecha como STRING en formato dd/mm/yyyy
    # NO convertir a datetime para evitar que Excel lo interprete mal
    if "fecha" in formatted.columns:
        # Si ya viene en formato correcto (dd/mm/yyyy), dejarlo así
        # Si viene en otro formato, normalizarlo
        def clean_fecha(val):
            if pd.isna(val) or val == "":
                return ""
            val_str = str(val)
            # Si ya tiene formato dd/mm/yyyy, dejarlo
            if "/" in val_str and len(val_str) == 10:
                return val_str
            # Si es datetime, convertir
            try:
                dt = pd.to_datetime(val, errors="coerce", dayfirst=True)
                if pd.notna(dt):
                    return dt.strftime("%d/%m/%Y")
            except:
                pass
            return val_str
        
        formatted["fecha"] = formatted["fecha"].apply(clean_fecha)

    formatted.to_excel(output_file, index=False, sheet_name="Transactions")
    highlight_problem_rows(str(output_file))


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    input_dir = project_root / "input"
    output_dir = project_root / "output"

    input_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    pdf_files = sorted(input_dir.rglob("*.pdf"))
    if not pdf_files:
        logger.warning("No PDF files found in %s", input_dir)
        print("No PDF files found. Add them to the input folder and run again.")
        return

    extractor = UniversalBankExtractor()
    processed = 0
    failed = 0

    for pdf_file in pdf_files:
        relative_path = pdf_file.relative_to(input_dir)
        logger.info("Processing %s", relative_path)
        try:
            # 🔹 Obtener resultado completo (dict con tables, metadata, etc.)
            result = extractor.extract_from_pdf(str(pdf_file), filename_hint=pdf_file.name)
            
            # 🔹 Extraer el DataFrame de las tablas
            tables = result.get("tables", [])
            if not tables or not isinstance(tables[0], pd.DataFrame):
                logger.warning("No data extracted from %s", relative_path)
                failed += 1
                continue
            
            dataframe = tables[0]
            
            if dataframe.empty:
                logger.warning("No data extracted from %s", relative_path)
                failed += 1
                continue

            destination_dir = output_dir / relative_path.parent
            destination_dir.mkdir(parents=True, exist_ok=True)
            output_file = destination_dir / f"{pdf_file.stem}.xlsx"
            format_excel_output(dataframe, output_file)

            observations = dataframe.get("observaciones", pd.Series(dtype=str))
            flagged = int(observations.apply(lambda value: bool(str(value).strip())).sum())
            logger.info("Saved %s rows (%s flagged) to %s", len(dataframe), flagged, output_file)
            processed += 1
        except Exception as exc:  # pragma: no cover - auxiliary script
            logger.error("Failed to process %s: %s", relative_path, exc, exc_info=True)
            failed += 1

    logger.info("Processing complete: %s ok, %s failed", processed, failed)
    print("Processing complete. Check the output directory for results.")


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled by user")
    except Exception as exc:
        logger.error("Unexpected error: %s", exc, exc_info=True)
        input("Unexpected error. Press Enter to exit...")